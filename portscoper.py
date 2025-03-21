#!/usr/bin/env python3

import xml.etree.ElementTree as ET
from typing import Dict, List, Optional, Set, DefaultDict
from dataclasses import dataclass
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
import argparse
import sys
import json
from rich.console import Console
from rich.table import Table
from datetime import datetime

@dataclass
class Port:
    number: int
    protocol: str
    state: str
    service: str
    version: Optional[str] = None
    scripts: Dict[str, str] = None

    def __post_init__(self):
        if self.scripts is None:
            self.scripts = {}

@dataclass
class Host:
    ip: str
    hostname: Optional[str]
    ports: List[Port]
    os_matches: List[str]
    status: str

class PortScoper:
    VERSION = "1.0.0"  # Add version constant at class level
    
    def __init__(self):
        self.console = Console()
        self.hosts: List[Host] = []
        self.unique_ports: DefaultDict[int, Set[str]] = defaultdict(set)
        self.enumeration_commands: Dict[str, List[str]] = {}
        self.scan_times: Dict[str, datetime] = {}  # Track scan times for each host

    def display_banner(self) -> None:
        """Display the tool's banner."""
        banner = f"""[bold cyan]
 ██████╗  ██████╗ ██████╗ ████████╗███████╗ ██████╗ ██████╗ ██████╗ ███████╗██████╗ 
 ██╔══██╗██╔═══██╗██╔══██╗╚══██╔══╝██╔════╝██╔════╝██╔═══██╗██╔══██╗██╔════╝██╔══██╗
 ██████╔╝██║   ██║██████╔╝   ██║   ███████╗██║     ██║   ██║██████╔╝█████╗  ██████╔╝
 ██╔═══╝ ██║   ██║██╔══██╗   ██║   ╚════██║██║     ██║   ██║██╔═══╝ ██╔══╝  ██╔══██╗
 ██║     ╚██████╔╝██║  ██║   ██║   ███████║╚██████╗╚██████╔╝██║     ███████╗██║  ██║
 ╚═╝      ╚═════╝ ╚═╝  ╚═╝   ╚═╝   ╚══════╝ ╚═════╝ ╚═════╝ ╚═╝     ╚══════╝╚═╝  ╚═╝[/bold cyan]"""
        
        self.console.print(banner)
        self.console.print(f"\n[dim]Version {self.VERSION} - https://github.com/bmodha/PortScoper[/dim]\n")

    def parse_nmap_xml(self, input_file: str, merge_strategy: str = 'union') -> None:
        """Parse Nmap XML output file and merge with existing results."""
        parser = NmapXMLParser(input_file)
        new_hosts = parser.parse()
        
        if not self.hosts:
            # First file, just store the results
            self.hosts = new_hosts
        else:
            # Merge with existing results based on strategy
            self._merge_results(new_hosts, merge_strategy)
        
        self._process_unique_ports()
        self._generate_enumeration_commands()

    def _merge_results(self, new_hosts: List[Host], strategy: str) -> None:
        """Merge new scan results with existing results based on strategy."""
        if strategy == 'union':
            self._merge_union(new_hosts)
        elif strategy == 'intersection':
            self._merge_intersection(new_hosts)
        elif strategy == 'latest':
            self._merge_latest(new_hosts)

    def _merge_union(self, new_hosts: List[Host]) -> None:
        """Merge by including all findings from all scans."""
        host_map = {host.ip: host for host in self.hosts}
        
        for new_host in new_hosts:
            if new_host.ip in host_map:
                # Merge ports for existing host
                existing_ports = {(p.number, p.protocol): p for p in host_map[new_host.ip].ports}
                for new_port in new_host.ports:
                    key = (new_port.number, new_port.protocol)
                    if key not in existing_ports:
                        host_map[new_host.ip].ports.append(new_port)
                    elif new_port.state == "open" and existing_ports[key].state != "open":
                        # Update port if it's now open
                        existing_ports[key] = new_port
            else:
                # Add new host
                self.hosts.append(new_host)

    def _merge_intersection(self, new_hosts: List[Host]) -> None:
        """Merge by keeping only findings present in all scans."""
        host_map = {host.ip: host for host in self.hosts}
        new_host_map = {host.ip: host for host in new_hosts}
        
        # Keep only hosts present in both scans
        common_ips = set(host_map.keys()) & set(new_host_map.keys())
        self.hosts = [host for host in self.hosts if host.ip in common_ips]
        
        # For each remaining host, keep only ports present in both scans
        for host in self.hosts:
            existing_ports = {(p.number, p.protocol): p for p in host.ports}
            new_ports = {(p.number, p.protocol): p for p in new_host_map[host.ip].ports}
            
            # Keep only ports present in both scans
            common_ports = set(existing_ports.keys()) & set(new_ports.keys())
            host.ports = [p for p in host.ports if (p.number, p.protocol) in common_ports]

    def _merge_latest(self, new_hosts: List[Host]) -> None:
        """Merge by keeping the most recent scan data for each host."""
        host_map = {host.ip: host for host in self.hosts}
        
        for new_host in new_hosts:
            # Always use the latest scan data for each host
            if new_host.ip in host_map:
                host_map[new_host.ip] = new_host
            else:
                self.hosts.append(new_host)

    def _process_unique_ports(self) -> None:
        """Process and store unique ports and their services."""
        # Clear existing unique ports before reprocessing
        self.unique_ports.clear()
        
        for host in self.hosts:
            for port in host.ports:
                if port.state == "open":
                    self.unique_ports[port.number].add(port.service)

    def _generate_enumeration_commands(self) -> None:
        """Generate enumeration commands for each unique port/service combination."""
        for port, services in self.unique_ports.items():
            commands = []
            services_lower = {service.lower() for service in services}
            
            # Special handling for HTTP/HTTPS combination
            if 'http' in services_lower and 'https' in services_lower:
                commands.extend(self.common_commands["http_https_combined"])
            else:
                # Process services normally
                for service in services:
                    service_lower = service.lower()
                    if service_lower in self.common_commands:
                        commands.extend(self.common_commands[service_lower])
                    else:
                        commands.extend(self.common_commands["default"])
            
            # Remove any duplicate commands that might have been added
            commands = list(dict.fromkeys(commands))
            self.enumeration_commands[str(port)] = commands

    def generate_excel_report(self, output_file: str) -> None:
        """Generate Excel report with host information."""
        excel_writer = ExcelWriter(self.hosts, output_file)
        excel_writer.save()
        self.console.print(f"[green]Successfully created Excel report: {output_file}[/green]")

    def print_unique_ports_report(self) -> None:
        """Print a table of unique ports and their services."""
        table = Table(title="Unique Ports and Services")
        table.add_column("Port", justify="right", style="cyan")
        table.add_column("Services", style="magenta")

        for port in sorted(self.unique_ports.keys()):
            services = ", ".join(sorted(self.unique_ports[port]))
            table.add_row(str(port), services)

        self.console.print(table)

    def print_enumeration_commands(self) -> None:
        """Print enumeration commands for each port."""
        # Create a mapping of ports to hosts that have that port open
        port_to_hosts: DefaultDict[int, List[tuple[str, str]]] = defaultdict(list)
        for host in self.hosts:
            for port in host.ports:
                if port.state == "open":
                    # Store tuple of (ip, service) for each host
                    port_to_hosts[port.number].append((host.ip, port.service))

        # Sort ports numerically
        port_numbers = sorted(port_to_hosts.keys())

        # Print header
        self.console.print("\n[bold yellow]Enumeration Commands by Port Number:[/bold yellow]")
        
        for port in port_numbers:
            # Get all hosts and their services for this port
            hosts_and_services = port_to_hosts[port]
            
            # Print port header with all services found on this port
            unique_services = sorted(set(service for _, service in hosts_and_services))
            services_str = ", ".join(unique_services)
            self.console.print(f"\n[bold cyan]Port {port} - {services_str}[/bold cyan]")
            
            # Get the appropriate commands for each service
            commands = set()  # Use set to avoid duplicate commands
            for _, service in hosts_and_services:
                service_lower = service.lower()
                if service_lower in self.common_commands:
                    commands.update(self.common_commands[service_lower])
                else:
                    commands.update(self.common_commands["default"])
            
            # Print commands for each host
            for host_ip, _ in sorted(hosts_and_services):  # Sort by IP
                self.console.print(f"\n[bold blue]Target: {host_ip}[/bold blue]")
                # Print each command with the IP and port filled in
                for i, cmd_template in enumerate(sorted(commands), 1):
                    cmd = cmd_template.format(target=host_ip, port=port)
                    self.console.print(f"[blue]{i}. {cmd}[/blue]")

    def save_enumeration_commands(self, output_file: str) -> None:
        """Save enumeration commands to a file with actual IP addresses and ports."""
        # Create a mapping of ports to hosts that have that port open
        port_to_hosts: DefaultDict[int, List[tuple[str, str]]] = defaultdict(list)
        for host in self.hosts:
            for port in host.ports:
                if port.state == "open":
                    port_to_hosts[port.number].append((host.ip, port.service))

        # Convert the commands dict to a sorted list of dictionaries
        sorted_commands = []
        for port in sorted(map(int, self.enumeration_commands.keys())):
            port_str = str(port)
            hosts_and_services = port_to_hosts[port]
            
            # Get all hosts and their services for this port
            host_commands = []
            for host_ip, service in hosts_and_services:
                # Format commands for this specific host and port
                formatted_commands = [
                    cmd.format(target=host_ip, port=port)
                    for cmd in self.enumeration_commands[port_str]
                ]
                host_commands.append({
                    "ip": host_ip,
                    "service": service,
                    "commands": formatted_commands
                })

            sorted_commands.append({
                "port": port,
                "services": list(self.unique_ports[port]),
                "hosts": host_commands
            })

        with open(output_file, 'w') as f:
            json.dump({"ports": sorted_commands}, f, indent=2)
        self.console.print(f"[green]Saved enumeration commands to: {output_file}[/green]")

    @property
    def common_commands(self) -> Dict[str, List[str]]:
        """Define common enumeration commands for different services."""
        return {
            "http_https_combined": [
                "nmap -sV -p {port} -sC --script=http*,ssl* {target}",
                "gobuster dir -u http://{target}:{port} -w /usr/share/wordlists/dirbuster/directory-list-2.3-medium.txt -t 50",
                "gobuster dir -u https://{target}:{port} -w /usr/share/wordlists/dirbuster/directory-list-2.3-medium.txt -t 50 -k",
                "gobuster vhost -u http://{target}:{port} -w /usr/share/seclists/Discovery/DNS/subdomains-top1million-5000.txt",
                "gobuster vhost -u https://{target}:{port} -w /usr/share/seclists/Discovery/DNS/subdomains-top1million-5000.txt -k",
                "nikto -h {target} -p {port} -C all",
                "nikto -h {target} -p {port} -ssl -C all",
                "sslscan --show-certificate --no-colour {target}:{port}",
                "testssl.sh --severity HIGH {target}:{port}",
                "whatweb -a 3 http://{target}:{port}",
                "whatweb -a 3 https://{target}:{port}",
                "curl -v -X OPTIONS http://{target}:{port}/",
                "curl -vk -X OPTIONS https://{target}:{port}/",
                "wfuzz -c -z file,/usr/share/wordlists/wfuzz/general/common.txt --hc 404 http://{target}:{port}/FUZZ",
                "wfuzz -c -z file,/usr/share/wordlists/wfuzz/general/common.txt --hc 404 https://{target}:{port}/FUZZ"
            ],
            "http": [
                "nmap -sV -p {port} -sC --script=http-enum,http-title,http-headers,http-methods,http-robots.txt,http-webdav-scan,http-vuln-* {target}",
                "gobuster dir -u http://{target}:{port} -w /usr/share/wordlists/dirbuster/directory-list-2.3-medium.txt -t 50",
                "gobuster vhost -u http://{target}:{port} -w /usr/share/seclists/Discovery/DNS/subdomains-top1million-5000.txt",
                "nikto -h {target} -p {port} -C all",
                "whatweb -a 3 http://{target}:{port}",
                "curl -v -X OPTIONS http://{target}:{port}/",
                "wfuzz -c -z file,/usr/share/wordlists/wfuzz/general/common.txt --hc 404 http://{target}:{port}/FUZZ"
            ],
            "https": [
                "nmap -sV -p {port} -sC --script=ssl-enum-ciphers,http-enum,http-methods,http-robots.txt,http-vuln-*,ssl-heartbleed,ssl-poodle,ssl-ccs-injection {target}",
                "gobuster dir -u https://{target}:{port} -w /usr/share/wordlists/dirbuster/directory-list-2.3-medium.txt -t 50 -k",
                "gobuster vhost -u https://{target}:{port} -w /usr/share/seclists/Discovery/DNS/subdomains-top1million-5000.txt -k",
                "nikto -h {target} -p {port} -ssl -C all",
                "sslscan --show-certificate --no-colour {target}:{port}",
                "testssl.sh --severity HIGH {target}:{port}",
                "whatweb -a 3 https://{target}:{port}",
                "curl -vk -X OPTIONS https://{target}:{port}/"
            ],
            "ssh": [
                "nmap -p {port} -sV -sC --script=ssh-auth-methods,ssh2-enum-algos,ssh-hostkey,ssh-publickey-acceptance {target}",
                "hydra -L /usr/share/wordlists/metasploit/unix_users.txt -P /usr/share/wordlists/rockyou.txt {target} ssh -s {port}",
                "searchsploit --colour -t OpenSSH",
                "ssh-audit {target}:{port}"
            ],
            "smb": [
                "nmap -p {port} -sV -sC --script=smb-*,msrpc-enum {target}",
                "enum4linux -a {target}",
                "smbclient -L //{target} -N",
                "smbmap -H {target}",
                "crackmapexec smb {target}",
                "nmap -p {port} --script=smb-vuln* {target}"
            ],
            "ftp": [
                "telnet -vn {target} {port} - Try anonymous login via USER anonymous and password anonymous@example.com",
                "ftp {target} {port} - Try anonymous login via Name: anonymous and Password: blank or anonymous@example.com",
                'msfconsole -x "use auxiliary/scanner/ftp/ftp_version; set RHOSTS {target}; run"',
                'msfconsole -x "use auxiliary/scanner/ftp/ftp_anonymous; set RHOSTS {target}; run"',
                'msfconsole -x "use auxiliary/scanner/ftp/ftp_login; set RHOSTS {target}; set USER_FILE users.txt; set PASS_FILE passwords.txt; run',
                "nmap -p {port} -sV -sC --script=ftp-anon,ftp-bounce {target} - anonymous login, and scan using ftp bounce method",
                "hydra -l admin -p /usr/share/wordlists/rockyou.txt {target} ftp -s {port} - brute force admin on non-standard port",
                "hydra -L /usr/share/wordlists/metasploit/unix_users.txt -P /usr/share/wordlists/rockyou.txt {target} ftp -s {port} - brute force on non-standard port",
                "medusa -h {target} -U /usr/share/wordlists/metasploit/unix_users.txt -P /usr/share/wordlists/rockyou.txt -M ftp -n {port}"
            ],
            "mysql": [
                "mysql -h {target} -u root -p <password> - Try default credentials, blank password",
                "nmap -p {port} --script mysql-info,mysql-audit,mysql-empty-password {target}",
                "nmap -p {port} -sV -sC --script=mysql-audit,mysql-databases,mysql-dump-hashes,mysql-empty-password,mysql-enum,mysql-info,mysql-query,mysql-users,mysql-variables,mysql-vuln-cve2012-2122 {target}",
                'msfconsole -x "use auxiliary/scanner/mysql/mysql_version; set RHOSTS {target}; run"',
                'msfconsole -x "use auxiliary/scanner/mysql/mysql_login; set RHOSTS {target}; set USER_FILE users.txt; set PASS_FILE passwords.txt; run"',
                'msfconsole -x "use auxiliary/scanner/mysql/mysql_root_password; set RHOSTS {target}; run"',
                'msfconsole -x "use auxiliary/scanner/mysql/mysql_enum; set RHOSTS {target}; run"',
                "hydra -L /usr/share/wordlists/metasploit/unix_users.txt -P /usr/share/wordlists/rockyou.txt {target} mysql -s {port}",
                "mysqldump --help | grep ssl"
            ],
            "mssql": [
                'msfconsole -x "use auxiliary/scanner/mssql/mssql_version; set RHOSTS {target}; run"',
                'msfconsole -x "use auxiliary/scanner/mssql/mssql_login; set RHOSTS {target}; set USER_FILE users.txt; set PASS_FILE passwords.txt; run"',
                'msfconsole -x "use auxiliary/scanner/mssql/mssql_enum; set RHOSTS {target}; run"',
                "nmap -p {port} -sV -sC --script=ms-sql-info,ms-sql-config,ms-sql-empty-password,ms-sql-ntlm-info,ms-sql-tables {target}",
                "hydra -L /usr/share/wordlists/metasploit/unix_users.txt -P /usr/share/wordlists/rockyou.txt {target} mssql -s {port}",
                "sqsh -S {target}:{port}"
            ],
            "postgresql": [
                "psql -h {target} -U postgres",
                "psql -h {target} -U postgres -d <databaseName>",
                'psql -h {target} -U postgres -d mydatabase -c "SELECT version();"',
                "nmap -p {port} -sV -sC --script=pgsql-version {target}",
                "nmap -p 5432 --script=pgsql-brute --script-args userdb=users.txt,passdb=passwords.txt {target}",
                'msfconsole -x "use auxiliary/scanner/postgres/postgres_login; set RHOSTS {target}; set USER_FILE users.txt; set PASS_FILE passwords.txt; run"'
            ],
            "redis": [
                "nmap -p {port} -sV -sC --script=redis-info {target}",
                "redis-cli -h {target} -p {port} info"
            ],
            "memcached": [
                "nmap -p {port} -sV -sC --script=memcached-info {target}",
                "telnet {target} {port}",
                "echo 'stats' | nc {target} {port}"
            ],
            "mongodb": [
                "mongo {target}:{port}",
                "mongo -u '<username>' -p '<password>' {target}:{port}",
                "mongo -u '<username>' -p '<password>' {target}:{port}/mydatabase",
                'mongo "mongodb://<username>:<password>@{target}:{port}/mydatabase?authSource=admin&ssl=true"',
                "mongo -u 'your_username' -p 'your_password' --authenticationDatabase admin {target}:{port}",
                "nmap -p {port} -sV -sC --script=mongodb-info,mongodb-databases {target}",
                "mongo {target}:{port} --eval 'db.version()'"
            ],
            "ldap": [
                "nmap -p {port} -sV -sC --script=ldap-rootdse,ldap-search,ldap-brute {target}",
                "ldapsearch -x -h {target} -p {port} -s base",
                "ldapsearch -x -h {target} -p {port} -b 'dc=example,dc=com'",
                'msfconsole -x "use auxiliary/scanner/ldap/ldap_version; set RHOSTS {target}; run"',
                'msfconsole -x "use auxiliary/scanner/ldap/ldap_anonymous; set RHOSTS {target}; run"'
            ],
            "dns": [
                "nc -nv -u {target} {port}",
                "nmap -p {port} --script dns-nsid {target}",
                "nmap -p {port} -sV -sC --script=dns-zone-transfer,dns-srv-enum,dns-recursion {target}",
                "dig @{target} version.bind chaos txt",
                "dnsenum --dnsserver {target} --enum example.com"
            ],
            "smtp": [
                "telnet {target} {port}",
                'msfconsole -x "use auxiliary/scanner/smtp/smtp_version; set RHOSTS {target}; run"',
                'msfconsole -x "use auxiliary/scanner/smtp/smtp_banner; set RHOSTS {target}; run"',
                "nmap -p {port} -sV -sC --script=smtp-commands,smtp-enum-users,smtp-vuln-* {target}",
                "nmap -p {port} --script=smtp-brute --script-args userdb=users.txt,passdb=passwords.txt {target}",
                "nmap -p 25 --script=smtp-vuln-cve2011-1720 {target} - SMTP Email Verification",
                "swaks --to user@example.com --from test@test.com --server {target}:{port}"
            ],
            "pop3": [
                "nmap -p {port} -sV -sC --script=pop3-capabilities {target}",
                'msfconsole -x "use auxiliary/scanner/pop3/pop3_version; set RHOSTS {target}; run"',
                'msfconsole -x "use auxiliary/scanner/pop3/pop3_banner; set RHOSTS {target}; run"',
                "nmap -p {port} --script=pop3-brute --script-args userdb=users.txt,passdb=passwords.txt {target}",
            ],
            "imap": [
                "nmap -p {port} -sV -sC --script=imap-capabilities {target}",
                "nmap -p 143 --script=imap-search {target}",
                'msfconsole -x "use auxiliary/scanner/imap/imap_version; set RHOSTS {target}; run"',
                'msfconsole -x "use auxiliary/scanner/imap/imap_banner; set RHOSTS {target}; run"',
                'msfconsole -x "use auxiliary/scanner/imap/imap_anonymous; set RHOSTS {target}; run"',
                "nmap -p {port} --script=imap-brute --script-args userdb=users.txt,passdb=passwords.txt {target}"
            ],
            "snmp": [
                'msfconsole -x "use auxiliary/scanner/snmp/snmp_enum; set RHOSTS {target}; run"',
                "nmap -p {port} -sV -sC --script=snmp-info,snmp-sysdescr {target}",
                "nmap -p 161 --script=snmp-brute --script-args userdb=community_strings.txt {target}",
                "snmpwalk -v2c -c public {target}:{port}",
                "onesixtyone -c /usr/share/seclists/Discovery/SNMP/snmp.txt {target}"
            ],
            "rdp": [
                "nmap -p {port} -sV -sC --script=rdp-version,rdp-enum-encryption,rdp-vuln-ms12-020 {target}",
                "hydra -L /usr/share/wordlists/metasploit/unix_users.txt -P /usr/share/wordlists/rockyou.txt {target} rdp -s {port}",
                'msfconsole -x "use auxiliary/scanner/rdp/rdp_login; set RHOSTS {target}; set USER_FILE users.txt; set PASS_FILE passwords.txt; run"',
                "xfreerdp /v:{target}:{port} /cert-ignore"
            ],
            "vnc": [
                "nmap -p {port} -sV -sC --script=vnc-info,realvnc-auth-bypass {target}",
                'msfconsole -x "use auxiliary/scanner/vnc/vnc_version; set RHOSTS {target}; run"',
                'msfconsole -x "use auxiliary/scanner/vnc/vnc_login; set RHOSTS {target}; set USER_FILE users.txt; set PASS_FILE passwords.txt; run"'
            ],
            "oracle": [
                "sqlplus username/password@{target}:{port}/oracle_service_name_or_sid",
                "nmap -p {port} -sV -sC --script=oracle-sid-brute,oracle-tns-version {target}",
                'msfconsole -x "use auxiliary/scanner/oracle/oracle_info; set RHOSTS {target}; run"',
                "tnscmd10g version -h {target} -p {port}"
            ],
            "telnet": [
                "telnet {target} {port}",
                "nmap -p 23 --script=telnet-ntlm-info {target}",
                'msfconsole -x "use auxiliary/scanner/telnet/telnet_login; set RHOSTS {target}; set USER_FILE users.txt; set PASS_FILE passwords.txt; run"',
                "nmap -p {port} -sV -sC --script=telnet-encryption,telnet-brute {target}"
            ],
            "docker": [
                "nmap -p {port} -sV -sC --script=docker-version,docker-api-info {target}",
                "curl -s -X GET http://{target}:{port}/version",
                "curl -s -X GET http://{target}:{port}/info"
            ],
            "docker-registry": [
                "nmap -p {port} -sV -sC --script=docker-registry-list-repositories {target}",
                "curl -s -X GET http://{target}:{port}/v2/_catalog",
                "curl -s -X GET http://{target}:{port}/v2/tags/list"
            ],
            "kerberos": [
                "nmap -p {port} -sV -sC --script=krb5-enum-users {target}",
                "GetADUsers.py -dc-ip {target} example.com/",
                "GetNPUsers.py -dc-ip {target} -request example.com/"
            ],
            "default": [
                "nmap -sV -p {port} -sC --script=banner,version {target}",
                "amap -bqv {target} {port}",
                "curl -v telnet://{target}:{port}"
            ]
        }

class NmapXMLParser:
    def __init__(self, input_file: str):
        self.input_file = Path(input_file)
        self.hosts: List[Host] = []

    def parse(self) -> List[Host]:
        """Parse Nmap XML output file."""
        if not self.input_file.exists():
            raise FileNotFoundError(f"File {self.input_file} not found")

        tree = ET.parse(self.input_file)
        root = tree.getroot()
        
        for host_elem in root.findall('.//host'):
            # Get IP address
            addr_elem = host_elem.find('.//address[@addrtype="ipv4"]')
            if addr_elem is None:
                continue
            ip = addr_elem.get('addr')
            
            # Get hostname if available
            hostname_elem = host_elem.find('.//hostname')
            hostname = hostname_elem.get('name') if hostname_elem is not None else None
            
            # Get status
            status = host_elem.find('status').get('state')
            
            # Get ports
            ports = []
            for port_elem in host_elem.findall('.//port'):
                port_number = int(port_elem.get('portid'))
                protocol = port_elem.get('protocol')
                
                state_elem = port_elem.find('state')
                state = state_elem.get('state') if state_elem is not None else 'unknown'
                
                service_elem = port_elem.find('service')
                service = service_elem.get('name') if service_elem is not None else 'unknown'
                version = service_elem.get('product') if service_elem is not None else None
                
                # Get script outputs
                scripts = {}
                for script_elem in port_elem.findall('script'):
                    script_id = script_elem.get('id')
                    script_output = script_elem.get('output')
                    scripts[script_id] = script_output
                
                ports.append(Port(port_number, protocol, state, service, version, scripts))
            
            # Get OS matches
            os_matches = []
            for os_elem in host_elem.findall('.//osmatch'):
                os_matches.append(os_elem.get('name'))
            
            self.hosts.append(Host(ip, hostname, ports, os_matches, status))
        
        return self.hosts

class ExcelWriter:
    def __init__(self, hosts: List[Host], output_file: str, organization: str = 'subnet'):
        self.hosts = hosts
        self.output_file = output_file
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)
        self.organization = organization  # 'subnet' or 'ip'
        self.subnet_groups = self._group_hosts_by_subnet() if organization == 'subnet' else None

    def _group_hosts_by_subnet(self) -> Dict[str, List[Host]]:
        """Group hosts by their /24 subnet."""
        subnet_groups = defaultdict(list)
        for host in self.hosts:
            # Get the /24 subnet by taking the first three octets
            subnet = '.'.join(host.ip.split('.')[:3]) + '.0/24'
            subnet_groups[subnet].append(host)
        return dict(subnet_groups)

    def create_summary_sheet(self) -> None:
        """Create a summary sheet with network overview."""
        ws = self.workbook.create_sheet("Summary")
        
        # Set column widths
        ws.column_dimensions['A'].width = 20  # Subnet
        ws.column_dimensions['B'].width = 15  # Host Count
        ws.column_dimensions['C'].width = 30  # Unique Services
        ws.column_dimensions['D'].width = 20  # Open Ports
        ws.column_dimensions['E'].width = 40  # Key Findings
        
        # Create headers
        headers = ['Subnet', 'Host Count', 'Unique Services', 'Open Ports', 'Key Findings']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Add subnet information
        row = 2
        for subnet, hosts in self.subnet_groups.items():
            # Collect subnet statistics
            host_count = len(hosts)
            unique_services = set()
            open_ports = set()
            for host in hosts:
                for port in host.ports:
                    if port.state == "open":
                        open_ports.add(port.number)
                        unique_services.add(port.service)
            
            # Create hyperlink to subnet sheet
            subnet_sheet_name = subnet.replace('.', '_').replace('/', '_')
            cell = ws.cell(row=row, column=1)
            cell.value = subnet
            cell.hyperlink = f"#{subnet_sheet_name}!A1"
            cell.font = Font(color="0563C1", underline="single")
            
            # Add statistics
            ws.cell(row=row, column=2).value = host_count
            ws.cell(row=row, column=3).value = ', '.join(sorted(unique_services))
            ws.cell(row=row, column=4).value = ', '.join(map(str, sorted(open_ports)))
            
            # Add key findings (e.g., critical services, potential vulnerabilities)
            key_findings = []
            critical_services = {'http', 'https', 'ssh', 'smb', 'mssql', 'mysql'}
            found_critical = critical_services.intersection(unique_services)
            if found_critical:
                key_findings.append(f"Critical services: {', '.join(sorted(found_critical))}")
            
            ws.cell(row=row, column=5).value = '; '.join(key_findings)
            
            row += 1

    def create_subnet_sheet(self, subnet: str, hosts: List[Host]) -> None:
        """Create a worksheet for a subnet."""
        sheet_name = subnet.replace('.', '_').replace('/', '_')
        ws = self.workbook.create_sheet(sheet_name)

        # Set column widths
        ws.column_dimensions['A'].width = 15  # IP
        ws.column_dimensions['B'].width = 20  # Hostname
        ws.column_dimensions['C'].width = 15  # Port
        ws.column_dimensions['D'].width = 15  # Protocol
        ws.column_dimensions['E'].width = 15  # State
        ws.column_dimensions['F'].width = 20  # Service
        ws.column_dimensions['G'].width = 30  # Version
        ws.column_dimensions['H'].width = 40  # Notes

        # Create headers
        headers = ['IP', 'Hostname', 'Port', 'Protocol', 'State', 'Service', 'Version', 'Notes']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Add subnet information header
        subnet_info = f"Subnet: {subnet} - Total Hosts: {len(hosts)}"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        cell = ws.cell(row=2, column=1)
        cell.value = subnet_info
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')

        # Add host information
        row = 3
        for host in sorted(hosts, key=lambda x: tuple(map(int, x.ip.split('.')))):
            for port in sorted(host.ports, key=lambda x: x.number):
                ws.cell(row=row, column=1).value = host.ip
                ws.cell(row=row, column=2).value = host.hostname if host.hostname else 'N/A'
                ws.cell(row=row, column=3).value = port.number
                ws.cell(row=row, column=4).value = port.protocol
                ws.cell(row=row, column=5).value = port.state
                ws.cell(row=row, column=6).value = port.service
                ws.cell(row=row, column=7).value = port.version if port.version else 'N/A'
                
                # Add script output to notes
                if port.scripts:
                    notes = []
                    for script_id, output in port.scripts.items():
                        notes.append(f"{script_id}: {output}")
                    ws.cell(row=row, column=8).value = '\n'.join(notes)
                
                row += 1

            # Add OS information if available
            if host.os_matches:
                ws.cell(row=row, column=1).value = host.ip
                ws.cell(row=row, column=2).value = "OS Detection"
                # Fix: Merge cells for OS information and put it in the Notes column
                ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
                ws.cell(row=row, column=6).value = ', '.join(host.os_matches)
                row += 1

            # Add a blank row between hosts
            row += 1

    def create_ip_sheet(self, host: Host) -> None:
        """Create a worksheet for a single IP."""
        sheet_name = host.ip.replace('.', '_')
        ws = self.workbook.create_sheet(sheet_name)

        # Set column widths
        ws.column_dimensions['A'].width = 15  # Port
        ws.column_dimensions['B'].width = 15  # Protocol
        ws.column_dimensions['C'].width = 15  # State
        ws.column_dimensions['D'].width = 20  # Service
        ws.column_dimensions['E'].width = 30  # Version
        ws.column_dimensions['F'].width = 40  # Notes

        # Create headers
        headers = ['Port', 'Protocol', 'State', 'Service', 'Version', 'Notes']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Add host information header
        host_info = f"IP: {host.ip} - Hostname: {host.hostname if host.hostname else 'N/A'}"
        if host.os_matches:
            host_info += f"\nOS: {', '.join(host.os_matches)}"
        
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        cell = ws.cell(row=2, column=1)
        cell.value = host_info
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')

        # Add port information
        row = 3
        for port in sorted(host.ports, key=lambda x: x.number):
            ws.cell(row=row, column=1).value = port.number
            ws.cell(row=row, column=2).value = port.protocol
            ws.cell(row=row, column=3).value = port.state
            ws.cell(row=row, column=4).value = port.service
            ws.cell(row=row, column=5).value = port.version if port.version else 'N/A'
            
            if port.scripts:
                notes = []
                for script_id, output in port.scripts.items():
                    notes.append(f"{script_id}: {output}")
                ws.cell(row=row, column=6).value = '\n'.join(notes)
            
            row += 1

    def create_ip_summary_sheet(self) -> None:
        """Create a summary sheet for IP-based organization."""
        ws = self.workbook.create_sheet("Summary")
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # IP
        ws.column_dimensions['B'].width = 20  # Hostname
        ws.column_dimensions['C'].width = 15  # Open Ports
        ws.column_dimensions['D'].width = 30  # Services
        ws.column_dimensions['E'].width = 40  # Key Findings
        
        # Create headers
        headers = ['IP', 'Hostname', 'Open Ports', 'Services', 'Key Findings']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Add host information
        row = 2
        for host in sorted(self.hosts, key=lambda x: tuple(map(int, x.ip.split('.')))):
            # Create hyperlink to host sheet
            sheet_name = host.ip.replace('.', '_')
            cell = ws.cell(row=row, column=1)
            cell.value = host.ip
            cell.hyperlink = f"#{sheet_name}!A1"
            cell.font = Font(color="0563C1", underline="single")
            
            # Add host details
            ws.cell(row=row, column=2).value = host.hostname if host.hostname else 'N/A'
            
            # Collect open ports and services
            open_ports = []
            services = set()
            for port in host.ports:
                if port.state == "open":
                    open_ports.append(str(port.number))
                    services.add(port.service)
            
            ws.cell(row=row, column=3).value = ', '.join(sorted(open_ports))
            ws.cell(row=row, column=4).value = ', '.join(sorted(services))
            
            # Add key findings
            key_findings = []
            critical_services = {'http', 'https', 'ssh', 'smb', 'mssql', 'mysql'}
            found_critical = critical_services.intersection(services)
            if found_critical:
                key_findings.append(f"Critical services: {', '.join(sorted(found_critical))}")
            if host.os_matches:
                key_findings.append(f"OS: {', '.join(host.os_matches)}")
            
            ws.cell(row=row, column=5).value = '; '.join(key_findings)
            
            row += 1

    def save(self) -> None:
        """Save the Excel workbook."""
        if self.organization == 'subnet':
            # Create subnet-based organization
            self.create_summary_sheet()
            for subnet, hosts in self.subnet_groups.items():
                self.create_subnet_sheet(subnet, hosts)
        else:
            # Create IP-based organization
            self.create_ip_summary_sheet()
            for host in self.hosts:
                self.create_ip_sheet(host)
        
        self.workbook.save(self.output_file)

def main():
    parser = argparse.ArgumentParser(
        description='PortScoper - A comprehensive Nmap output analyzer and enumeration helper',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage (uses default output filenames)
  python portscoper.py scan.xml

  # Multiple input files
  python portscoper.py -i scan1.xml scan2.xml scan3.xml

  # Specify custom output files
  python portscoper.py -i scan1.xml scan2.xml -o custom_report.xlsx -c custom_commands.json

  # Choose organization mode (subnet or ip)
  python portscoper.py -i scan.xml --organization ip

Note:
  The input files must be in Nmap XML format. Generate them using:
  nmap -sV -sC -O <target> -oX scan.xml
""")
    
    parser.add_argument(
        '-i', '--input-files',
        nargs='+',
        help='One or more input Nmap XML files (generated using -oX flag)',
        required=True
    )
    parser.add_argument(
        '-o', '--output',
        help='Output Excel file path (default: %(default)s)',
        default='portscoper_report.xlsx'
    )
    parser.add_argument(
        '-c', '--commands',
        help='Output file path for enumeration commands in JSON format (default: %(default)s)',
        default='enumeration_commands.json'
    )
    parser.add_argument(
        '--merge-strategy',
        choices=['union', 'intersection', 'latest'],
        default='union',
        help='Strategy to merge multiple scan results: union (all findings), intersection (common findings), or latest (most recent scan data for each host)'
    )
    parser.add_argument(
        '--organization',
        choices=['subnet', 'ip'],
        default='subnet',
        help='How to organize the Excel sheets: by subnet or by individual IP (default: subnet)'
    )
    
    args = parser.parse_args()
    
    try:
        # Initialize PortScoper
        scoper = PortScoper()
        
        # Display banner
        scoper.display_banner()
        
        # Parse multiple nmap outputs
        scoper.console.print("[yellow]Parsing Nmap XML outputs...[/yellow]")
        for input_file in args.input_files:
            scoper.console.print(f"[yellow]Processing {input_file}...[/yellow]")
            scoper.parse_nmap_xml(input_file, merge_strategy=args.merge_strategy)
        
        # Generate Excel report
        scoper.console.print("[yellow]Generating Excel report...[/yellow]")
        scoper.generate_excel_report(args.output)
        
        # Print unique ports report
        scoper.console.print("\n[yellow]Unique Ports and Services:[/yellow]")
        scoper.print_unique_ports_report()
        
        # Print and save enumeration commands
        scoper.console.print("\n[yellow]Enumeration Commands:[/yellow]")
        scoper.print_enumeration_commands()
        scoper.save_enumeration_commands(args.commands)
        
    except Exception as e:
        scoper.console.print(f"[red]Error: {str(e)}[/red]", style="bold red")
        sys.exit(1)

if __name__ == "__main__":
    main() 
